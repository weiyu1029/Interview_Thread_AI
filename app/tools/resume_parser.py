from __future__ import annotations

import csv
import email
import json
import re
import zipfile
from io import BytesIO
from xml.etree import ElementTree

MAX_RESUME_BYTES = 12 * 1024 * 1024
SUPPORTED_EXTENSIONS = (
    "pdf", "docx", "odt", "rtf", "txt", "md", "html", "htm", "csv", "json", "xlsx",
    "pptx", "epub", "eml", "xml", "yaml", "yml", "log", "tex",
)
MAX_ARCHIVE_MEMBERS = 2_000
MAX_EXPANDED_BYTES = 60 * 1024 * 1024


def _safe_zip(content: bytes) -> zipfile.ZipFile:
    archive = zipfile.ZipFile(BytesIO(content))
    members = archive.infolist()
    if len(members) > MAX_ARCHIVE_MEMBERS or sum(item.file_size for item in members) > MAX_EXPANDED_BYTES:
        archive.close()
        raise ValueError("Compressed document expands beyond the safe processing limit.")
    return archive


def _xml_text(content: bytes) -> str:
    root = ElementTree.fromstring(content)
    return "\n".join(piece.strip() for piece in root.itertext() if piece.strip())


def extract_resume_text(file_name: str, content: bytes) -> str:
    """Extract resume text from an in-memory public-web upload.

    The file is never written to disk. Callers should keep the documented 12 MB
    limit so a public deployment cannot be used for unbounded file processing.
    """
    if len(content) > MAX_RESUME_BYTES:
        raise ValueError("Resume files must be 12 MB or smaller.")

    suffix = file_name.lower().rsplit(".", maxsplit=1)[-1] if "." in file_name else ""
    if suffix == "pdf":
        from pypdf import PdfReader

        reader = PdfReader(BytesIO(content))
        return "\n".join((page.extract_text() or "") for page in reader.pages).strip()

    if suffix == "docx":
        from docx import Document

        document = Document(BytesIO(content))
        return "\n".join(
            paragraph.text
            for paragraph in document.paragraphs
            if paragraph.text.strip()
        ).strip()

    if suffix == "odt":
        from odf import teletype
        from odf.opendocument import load
        from odf.text import P

        document = load(BytesIO(content))
        return "\n".join(teletype.extractText(paragraph) for paragraph in document.getElementsByType(P)).strip()

    if suffix == "rtf":
        from striprtf.striprtf import rtf_to_text

        return rtf_to_text(content.decode("utf-8", errors="replace")).strip()

    if suffix in {"html", "htm"}:
        from bs4 import BeautifulSoup

        return BeautifulSoup(content, "html.parser").get_text("\n", strip=True)

    if suffix == "csv":
        decoded = content.decode("utf-8-sig", errors="replace")
        return "\n".join(" | ".join(cell.strip() for cell in row if cell.strip()) for row in csv.reader(decoded.splitlines())).strip()

    if suffix == "json":
        payload = json.loads(content.decode("utf-8-sig", errors="replace"))

        def flatten(value: object) -> list[str]:
            if isinstance(value, dict):
                return [f"{key}: {'; '.join(flatten(item))}" for key, item in value.items()]
            if isinstance(value, list):
                return [piece for item in value for piece in flatten(item)]
            return [str(value)]

        return "\n".join(flatten(payload)).strip()

    if suffix == "xlsx":
        from openpyxl import load_workbook

        workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
        lines: list[str] = []
        for sheet in workbook.worksheets:
            lines.append(f"Sheet: {sheet.title}")
            for row in sheet.iter_rows(values_only=True):
                values = [str(value).strip() for value in row if value not in (None, "")]
                if values:
                    lines.append(" | ".join(values))
        return "\n".join(lines).strip()

    if suffix == "pptx":
        with _safe_zip(content) as archive:
            slides = sorted(name for name in archive.namelist() if re.fullmatch(r"ppt/slides/slide\d+\.xml", name))
            return "\n".join(_xml_text(archive.read(name)) for name in slides).strip()

    if suffix == "epub":
        from bs4 import BeautifulSoup

        with _safe_zip(content) as archive:
            pages = sorted(name for name in archive.namelist() if name.lower().endswith((".xhtml", ".html", ".htm")))
            return "\n".join(BeautifulSoup(archive.read(name), "html.parser").get_text("\n", strip=True) for name in pages).strip()

    if suffix == "eml":
        message = email.message_from_bytes(content)
        chunks = [f"Subject: {message.get('subject', '')}"]
        for part in message.walk():
            if part.get_content_type() == "text/plain" and part.get_content_disposition() != "attachment":
                payload = part.get_payload(decode=True) or b""
                chunks.append(payload.decode(part.get_content_charset() or "utf-8", errors="replace"))
        return "\n".join(chunks).strip()

    if suffix == "xml":
        return _xml_text(content).strip()

    if suffix in {"txt", "md", "yaml", "yml", "log", "tex"}:
        return content.decode("utf-8", errors="replace").strip()

    supported = ", ".join(item.upper() for item in SUPPORTED_EXTENSIONS)
    raise ValueError(f"Supported document formats: {supported}.")
